# -*- coding:utf-8 -*-  
"""
Create on 16/12/21
Author xiaoyy
"""
from openpyxl import Workbook
import pymysql
import json
import six
from pyspider.libs import utils
from demo.common.constants import SEARCH_LIST2
from demo.common.colors import font1

title_row_all = ['名称', '类别', '营养素', '热量(大卡)', '碳水化合物(克)', '脂肪(克)', '蛋白质(克)', '纤维素(克)', '维生素A(微克)', '维生素C(毫克)',
                 '维生素E(毫克)', '胡萝卜素(微克)', '硫胺素(毫克)', '核黄素(毫克)', '烟酸(毫克)', '胆固醇(毫克)', '镁(毫克)', '钙(毫克)', '铁(毫克)', '锌(毫克)',
                 '铜(毫克)', '锰(毫克)', '钾(毫克)', '磷(毫克)', '钠(毫克)', '硒(微克)']
# title_row_all = ['名称', '类别', '别名', '营养素', '热量(大卡)', '碳水化合物(克)', '脂肪(克)', '蛋白质(克)', '纤维素(克)', '维生素A(微克)', '维生素C(毫克)',
#                  '维生素E(毫克)', '胡萝卜素(微克)', '硫胺素(毫克)', '核黄素(毫克)', '烟酸(毫克)', '胆固醇(毫克)', '镁(毫克)', '钙(毫克)', '铁(毫克)', '锌(毫克)',
#                  '铜(毫克)', '锰(毫克)', '钾(毫克)', '磷(毫克)', '钠(毫克)', '硒(微克)']
title_row = ['营养素', '热量(大卡)', '碳水化合物(克)', '脂肪(克)', '蛋白质(克)', '纤维素(克)', '维生素A(微克)', '维生素C(毫克)', '维生素E(毫克)', '胡萝卜素(微克)',
             '硫胺素(毫克)', '核黄素(毫克)', '烟酸(毫克)', '胆固醇(毫克)', '镁(毫克)', '钙(毫克)', '铁(毫克)', '锌(毫克)', '铜(毫克)', '锰(毫克)', '钾(毫克)',
             '磷(毫克)', '钠(毫克)', '硒(微克)']

type_row = ['雀巢']


def search_blob_demo():
    # 连接配置信息
    mysql_config = {
        'host': '192.168.1.244',
        'port': 3306,
        'user': 'root',
        'password': 'root',
        'db': 'resultdb',
        'charset': 'utf8',
        'cursorclass': pymysql.cursors.DictCursor,
    }
    # 创建连接
    connection = pymysql.connect(**mysql_config)
    # 执行sql语句
    try:
        with connection.cursor() as cursor:
            # 执行sql语句，进行查询
            sql = 'select * from boohee6'
            # 获取查询结果
            cursor.execute(sql)
            data = cursor.fetchall()
            wb = Workbook()
            # 激活 worksheet
            sheet = wb.create_sheet('薄荷网', 0)
            # 可以附加行，从第一列开始附加
            sheet.append(title_row_all)
            num = 1
            for j in range(len(SEARCH_LIST2)):
                search_key = SEARCH_LIST2[j]
                for i in range(len(data)):
                    for key, value in list(six.iteritems(data[i])):
                        if isinstance(value, (bytearray, six.binary_type)):
                            # data[key] = value.decode('utf8')
                            # 作用同上
                            data[i][key] = utils.text(value)
                    if 'result' in data[i]:
                        result = data[i]['result']
                        # 判断result类型为str
                        if isinstance(result, str):
                            data[i]['result'] = json.loads(data[i]['result'])
                    # 判断search_key 精准匹配name
                    if search_key != data[i]['result']['name']:
                        continue
                    else:
                        # 过滤type_row数据
                        if data[i]['result']['type'] in type_row:
                            continue
                        else:
                            search_key_title = []
                            search_key_title.append(search_key)
                            sheet.append(search_key_title)
                            num += 1
                            # 获取title行
                            title_cell = sheet.cell(None, num, 1)
                            title_cell.font = font1
                            # merge cells from num row
                            sheet.merge_cells(None, num, 1, num, 3)

                            contents = data[i]['result']['contents'].split('>>')[0].strip().split(' ')[2:]
                            content_all = []
                            for s in range(len(title_row)):
                                if title_row[s] in contents:
                                    for z in range(len(contents)):
                                        if title_row[s] == contents[z]:
                                            content_all.append('' if contents[z + 1] == '一' else contents[z + 1])
                                else:
                                    content_all.append('')
                            content_all.insert(0, data[i]['result']['name'])
                            content_all.insert(1, data[i]['result']['type'])
                            sheet.append(content_all)
                            num += 1
            # 保存文件
            wb.save("薄荷网食物data-precise1.xlsx")

        # 没有设置默认自动提交，需要主动提交，以保存所执行的语句
        connection.commit()
    finally:
        connection.close()


search_blob_demo()
